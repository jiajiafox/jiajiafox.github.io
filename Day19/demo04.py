from openpyxl import Workbook
import openpyxl
# 建立檔案
fileName="README.xlsx"
# wb=Workbook()   
# wb.save(fileName)

# 讀取
wb=openpyxl.load_workbook(fileName)

# 設定工作表
s1=wb['Sheet']

s1['A1'].value='java'   #  ['A1'] 代表A列第一個欄位
s1['A2'].value='python'
s1['A3'].value='php'
s1['A4'].value='C#'
s1.cell(1,2).value=30   # cell(1,2) 代表第一行第二個欄位
s1.cell(2,2).value=5
s1.cell(3,2).value=20
s1.cell(4,2).value=20

# 存入Excel
wb.save(fileName)
