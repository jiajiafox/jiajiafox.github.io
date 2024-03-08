from openpyxl import Workbook
import openpyxl


fileName="2.xlsx"
# wb=Workbook()   # 閱讀
# wb.save(fileName)

# 開啟excel
wb=openpyxl.load_workbook(fileName) 
s1=wb['工作表1']  # 讀取 sheet工作表
s2=wb.active

print(s1.title,s1.max_row,s1.max_column)